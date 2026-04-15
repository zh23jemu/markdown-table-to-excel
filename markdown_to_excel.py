from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


SEPARATOR_RE = re.compile(r"^\s*\|?(?:\s*:?-{3,}:?\s*\|)+\s*:?-{3,}:?\s*\|?\s*$")
BR_TAG_RE = re.compile(r"<br\s*/?>", re.IGNORECASE)
AGE_TOKEN_RE = re.compile(r"^(?:約)?[零一二三四五六七八九十百千廿卅卌半元兩二三四五六七八九十\d]+(?:歲|月|日|具)?$")
AGE_PREFIX_RE = re.compile(r"^((?:約)?(?:[零一二三四五六七八九十百千廿卅卌半元兩二三四五六七八九十\d]*歲|[零一二三四五六七八九十百千廿卅卌半元兩二三四五六七八九十\d]+(?:月|日|具)))(.+)$")
DATE_LIKE_RE = re.compile(
    r"[一二三四五六七八九十廿卅卌元\d]+月|[一二三四五六七八九十廿卅卌元\d]+日|又月又日|初[一二三四五六七八九十\d]|[Xx×✕]\s*月\s*[Xx×✕]\s*日"
)
ADDRESS_KEYWORDS = (
    "方便所",
    "福音堂",
    "福音医院",
    "福音醫院",
    "醫院",
    "医院",
    "街",
    "里",
    "巷",
    "路",
    "橋",
    "桥",
    "碼頭",
    "码头",
    "祠",
    "院",
    "棧",
    "栈",
    "埠頭",
    "碼頭水",
    "飯店",
    "坪",
    "池",
    "旗杆",
)
ORIGIN_SUFFIXES = ("縣", "县", "州", "鄉", "乡", "埠", "門", "门", "海", "浦", "寧", "宁", "甯", "陽", "阳", "溪", "島", "岛", "山", "岡", "冈", "口")


@dataclass
class TableBlock:
    title: str
    table_number: int
    headers: list[str]
    rows: list[list[str]]


HEADER_MERGE_GROUPS = (
    ("姓", "名"),
    ("性", "別"),
    ("年", "齡"),
    ("籍", "貫"),
    ("住", "址"),
    ("病", "狀"),
    ("認", "家"),
    ("死", "亡", "日", "期"),
    ("附", "記"),
    ("棺木", "仙衣", "類別"),
)

MATRIX_TRANSPOSE_ROW_HEADERS = (
    "籍貫",
    "住址",
    "病狀",
    "死亡日期",
    "墓地隴名",
    "墓地號數",
    "棺木類別",
    "關係人",
)

COFFIN_SUFFIXES = (
    "衣棺",
    "厚仔",
    "幫工",
    "帮工",
    "棺木",
    "棺工",
    "厚衣",
    "金箱",
    "金口相",
    "木",
    "棺",
    "工",
    "仔",
)


def normalize_cell(text: str) -> str:
    text = re.sub(r"\*\*(.*?)\*\*", r"\1", text)
    text = re.sub(r"\*(.*?)\*", r"\1", text)
    text = re.sub(r"__(.*?)__", r"\1", text)
    text = re.sub(r"_(.*?)_", r"\1", text)
    return text.strip()


def split_markdown_row(line: str) -> list[str]:
    stripped = line.strip()
    if stripped.startswith("|"):
        stripped = stripped[1:]
    if stripped.endswith("|"):
        stripped = stripped[:-1]
    return [normalize_cell(part) for part in stripped.split("|")]


def is_table_row(line: str) -> bool:
    stripped = line.strip()
    return stripped.startswith("|") and stripped.endswith("|") and "|" in stripped[1:-1]


def is_separator_line(line: str) -> bool:
    stripped = line.strip()
    return stripped == "---" or stripped == "***" or bool(SEPARATOR_RE.match(stripped))


def extract_title(candidate_lines: Iterable[str]) -> str:
    fallback_title = ""

    for raw_line in reversed(list(candidate_lines)):
        line = raw_line.strip()
        if not line or is_separator_line(line):
            continue
        if line.startswith("```"):
            continue
        if line.startswith("#"):
            return line.lstrip("#").strip()
        if not fallback_title:
            fallback_title = line

    return fallback_title


def parse_markdown_tables(path: str | Path) -> list[TableBlock]:
    file_path = Path(path)
    lines = file_path.read_text(encoding="utf-8").splitlines()
    tables: list[TableBlock] = []
    recent_lines: list[str] = []
    index = 0

    while index < len(lines):
        line = lines[index]

        if is_table_row(line) and index + 1 < len(lines) and is_separator_line(lines[index + 1]):
            headers = split_markdown_row(line)
            title = extract_title(recent_lines)
            index += 2
            rows: list[list[str]] = []

            while index < len(lines) and is_table_row(lines[index]):
                row = split_markdown_row(lines[index])
                if len(row) < len(headers):
                    row.extend([""] * (len(headers) - len(row)))
                elif len(row) > len(headers):
                    headers.extend([""] * (len(row) - len(headers)))
                    for existing_row in rows:
                        existing_row.extend([""] * (len(headers) - len(existing_row)))
                rows.append(row)
                index += 1

            headers, rows = merge_split_header_columns(headers, rows)
            if should_transpose_matrix_table(headers, rows):
                headers, rows = transpose_matrix_table(headers, rows)

            tables.append(
                TableBlock(
                    title=title or file_path.stem,
                    table_number=len(tables) + 1,
                    headers=headers,
                    rows=rows,
                )
            )
            recent_lines = []
            continue

        if line.strip() and not is_separator_line(line):
            recent_lines.append(line)
            if len(recent_lines) > 30:
                recent_lines.pop(0)
        elif not line.strip():
            recent_lines.append("")
            if len(recent_lines) > 30:
                recent_lines.pop(0)

        index += 1

    return tables


def merge_cells(values: list[str]) -> str:
    return "".join(value.strip() for value in values if value and value.strip())


def should_transpose_matrix_table(headers: list[str], rows: list[list[str]]) -> bool:
    normalized_headers = [header.strip() for header in headers]
    if any(normalized_headers):
        return False
    if len(headers) < 6 or len(rows) < 4:
        return False
    if len(headers) <= len(rows):
        return False

    first_row = rows[0] if rows else []
    non_empty_first_row = sum(1 for cell in first_row if cell.strip())
    multiline_cells = sum(1 for cell in first_row if "<br" in cell.lower())

    if non_empty_first_row < max(4, int(len(first_row) * 0.6)):
        return False
    if multiline_cells < max(3, int(len(first_row) * 0.5)):
        return False

    return True


def transpose_matrix_table(headers: list[str], rows: list[list[str]]) -> tuple[list[str], list[list[str]]]:
    max_columns = max([len(headers)] + [len(row) for row in rows])
    padded_rows = [row + [""] * (max_columns - len(row)) for row in rows]

    transposed_headers = ["姓名", "性別", "年齡"]
    for index in range(1, len(padded_rows)):
        if index - 1 < len(MATRIX_TRANSPOSE_ROW_HEADERS):
            transposed_headers.append(MATRIX_TRANSPOSE_ROW_HEADERS[index - 1])
        else:
            transposed_headers.append(f"字段{index + 3}")

    transposed_rows: list[list[str]] = []
    for column_index in range(max_columns):
        first_cell = padded_rows[0][column_index].strip()
        if not first_cell:
            continue

        segments = [segment.strip() for segment in BR_TAG_RE.split(first_cell) if segment.strip()]
        if len(segments) >= 3 and segments[1] in {"男", "女"} and is_age_like(segments[2]):
            row_values = [segments[0], segments[1], segments[2]]
            first_extra = "<br>".join(segments[3:]) if len(segments) > 3 else ""
        else:
            row_values = [first_cell, "", ""]
            first_extra = ""

        for row_index in range(1, len(padded_rows)):
            value = padded_rows[row_index][column_index].strip()
            if row_index == 1 and first_extra:
                value = f"{first_extra}<br>{value}" if value else first_extra
            row_values.append(value)

        if len(row_values) > 3:
            location_value = row_values[3].strip()
            origin_part = ""
            address_part = location_value
            for split_index in range(1, len(location_value)):
                left = location_value[:split_index].strip()
                right = location_value[split_index:].strip()
                if not left or not right:
                    continue
                if is_origin_like(left) and is_address_like(right):
                    origin_part = left
                    address_part = right
                    break
            row_values[3] = origin_part
            row_values.insert(4, address_part)

        transposed_rows.append(row_values)

    return transposed_headers, transposed_rows


def merge_split_header_columns(headers: list[str], rows: list[list[str]]) -> tuple[list[str], list[list[str]]]:
    normalized_headers = headers[:]
    normalized_rows = [row[:] for row in rows]
    column_index = 0

    while column_index < len(normalized_headers):
        matched_group: tuple[str, ...] | None = None
        for group in HEADER_MERGE_GROUPS:
            group_length = len(group)
            if tuple(normalized_headers[column_index:column_index + group_length]) == group:
                matched_group = group
                break

        if not matched_group:
            column_index += 1
            continue

        group_length = len(matched_group)
        merged_header = "".join(matched_group)
        normalized_headers = (
            normalized_headers[:column_index]
            + [merged_header]
            + normalized_headers[column_index + group_length:]
        )

        merged_rows: list[list[str]] = []
        for row in normalized_rows:
            padded_row = row + [""] * (column_index + group_length - len(row))
            merged_value = merge_cells(padded_row[column_index:column_index + group_length])
            merged_row = (
                padded_row[:column_index]
                + [merged_value]
                + padded_row[column_index + group_length:]
            )
            merged_rows.append(merged_row)
        normalized_rows = merged_rows
        column_index += 1

    return normalized_headers, normalized_rows


def split_br_delimited_cells(row: list[str], expected_width: int) -> list[str]:
    normalized_row = row + [""] * (expected_width - len(row))
    column_index = 0

    while column_index < expected_width:
        cell_value = normalized_row[column_index].strip()
        if "<br" not in cell_value.lower():
            column_index += 1
            continue

        segments = [segment.strip() for segment in BR_TAG_RE.split(cell_value)]
        if len(segments) <= 1 or any(not segment for segment in segments):
            column_index += 1
            continue

        right_empty_slots = 0
        probe_index = column_index + 1
        while probe_index < expected_width and not normalized_row[probe_index].strip():
            right_empty_slots += 1
            probe_index += 1

        if len(segments) != right_empty_slots + 1:
            column_index += 1
            continue

        for offset, segment in enumerate(segments):
            normalized_row[column_index + offset] = segment

        column_index += len(segments)

    return normalized_row


def find_header_index(headers: list[str], candidates: tuple[str, ...]) -> int | None:
    for index, header in enumerate(headers):
        normalized_header = header.replace(" ", "")
        if any(candidate in normalized_header for candidate in candidates):
            return index
    return None


def is_age_like(value: str) -> bool:
    stripped = value.strip()
    if not stripped:
        return False
    if "白骨" in stripped:
        return True
    return bool(AGE_TOKEN_RE.fullmatch(stripped))


def normalize_age_text(value: str) -> str:
    stripped = value.strip()
    if stripped.startswith("空"):
        candidate = stripped[1:].strip()
        if candidate and ("歲" in candidate or "岁" in candidate):
            return candidate
    return stripped


def is_age_marker_like(value: str) -> bool:
    stripped = value.strip()
    if not stripped:
        return False
    if stripped in {"歲", "岁"}:
        return True
    if is_age_like(stripped):
        return True
    return bool(re.fullmatch(r"(?:約)?[天夭零一二三四五六七八九十百千廿卅卌半元兩\d]*[歲岁]", stripped))


def split_gender_and_age(value: str) -> tuple[str, str] | None:
    stripped = normalize_age_text(value)
    if not stripped:
        return None

    match = re.fullmatch(r"([男女])(.+)", stripped)
    if not match:
        return None

    gender_part = match.group(1)
    age_part = match.group(2).strip()
    if not is_age_marker_like(age_part):
        return None

    return gender_part, age_part


def is_age_fragment(value: str) -> bool:
    stripped = value.strip()
    if not stripped:
        return False
    return bool(re.fullmatch(r"[零一二三四五六七八九十百千廿卅卌半元兩\d]+", stripped))


def split_age_and_origin(age_value: str) -> tuple[str, str] | None:
    stripped = normalize_age_text(age_value)
    if not stripped:
        return None

    match = AGE_PREFIX_RE.fullmatch(stripped)
    if not match:
        return None

    age_part = match.group(1).strip()
    origin_part = match.group(2).strip()
    if not age_part or not origin_part:
        return None

    return age_part, origin_part


def is_date_like(value: str) -> bool:
    stripped = value.strip()
    if not stripped:
        return False
    return bool(DATE_LIKE_RE.search(stripped))


def is_origin_like(value: str) -> bool:
    stripped = value.strip()
    if not stripped:
        return False
    if any(keyword in stripped for keyword in ADDRESS_KEYWORDS):
        return False
    return stripped.endswith(ORIGIN_SUFFIXES)


def is_address_like(value: str) -> bool:
    stripped = value.strip()
    if not stripped:
        return False
    return any(keyword in stripped for keyword in ADDRESS_KEYWORDS)


def extract_coffin_suffix(value: str) -> tuple[str, str]:
    stripped = value.strip()
    if not stripped:
        return "", ""

    for suffix in COFFIN_SUFFIXES:
        if stripped.endswith(suffix):
            return stripped[:-len(suffix)], suffix
    return stripped, ""


def split_compact_tail_fields(value: str) -> tuple[str, str, str, str] | None:
    stripped = value.strip()
    if not stripped:
        return None

    remainder, coffin = extract_coffin_suffix(stripped)
    if not coffin:
        return None

    family = ""
    death = ""
    grave = ""

    if remainder.startswith("又日"):
        death = "又月又日"
        grave = remainder[len("又日"):].strip()
    else:
        full_date_match = re.search(
            r"^(?P<family>.*?)(?P<death>[又元一二三四五六七八九十廿卅卌Xx×✕]*月[又元一二三四五六七八九十廿卅卌Xx×✕]*日)(?P<grave>.+)$",
            remainder,
        )
        if full_date_match:
            family = full_date_match.group("family").strip()
            death = full_date_match.group("death").strip()
            grave = full_date_match.group("grave").strip()

    if not death or not grave:
        return None

    if death in {"月又日", "月又日".replace(" ", "")}:
        death = "又月又日"
    elif death.startswith("月"):
        death = f"又{death}"

    if family and family[-1] in "一二三四五六七八九十廿卅卌元":
        if death and death[0] != family[-1]:
            death = family[-1] + death
            family = family[:-1].strip()

    return family, death, grave, coffin


def normalize_compact_obituary_row(headers: list[str], row: list[str]) -> list[str]:
    normalized_row = row[:]
    age_index = find_header_index(headers, ("年齡", "年龄"))
    origin_index = find_header_index(headers, ("籍貫", "籍贯"))
    address_index = find_header_index(headers, ("住址",))
    illness_index = find_header_index(headers, ("病狀", "病状"))
    family_index = find_header_index(headers, ("認家", "认家"))
    death_index = find_header_index(headers, ("死亡日期",))
    grave_index = find_header_index(headers, ("墓地號數", "墓地号数"))
    coffin_index = find_header_index(headers, ("棺木",))

    required_indexes = (age_index, origin_index, address_index, illness_index, family_index, death_index, grave_index, coffin_index)
    if any(index is None for index in required_indexes):
        return normalized_row

    max_index = max(index for index in required_indexes if index is not None)
    if len(normalized_row) <= max_index:
        normalized_row.extend([""] * (max_index + 1 - len(normalized_row)))

    if age_index is not None:
        normalized_row[age_index] = normalize_age_text(normalized_row[age_index])
        age_and_origin = split_age_and_origin(normalized_row[age_index])
        if age_and_origin:
            if normalized_row[origin_index].strip():
                normalized_row = shift_row_segment_right(normalized_row, origin_index)
            normalized_row[age_index], normalized_row[origin_index] = age_and_origin

    family_seed = ""
    origin_value = normalized_row[origin_index].strip()
    address_value = normalized_row[address_index].strip()
    illness_value = normalized_row[illness_index].strip()
    if address_value.endswith("月") and illness_value in {"又日", "月四"}:
        family_seed = address_value[:-1].strip()
        if origin_value and is_address_like(origin_value):
            normalized_row[address_index] = origin_value
            normalized_row[origin_index] = ""
        else:
            normalized_row[address_index] = ""

    tail_source = "".join(
        normalized_row[index].strip()
        for index in range(illness_index, coffin_index + 1)
        if normalized_row[index].strip()
    )
    compact_tail = split_compact_tail_fields(tail_source)
    if not compact_tail:
        return normalized_row

    family_value, death_value, grave_value, coffin_value = compact_tail
    normalized_row[illness_index] = ""
    normalized_row[family_index] = family_value or family_seed
    normalized_row[death_index] = death_value
    normalized_row[grave_index] = grave_value
    normalized_row[coffin_index] = coffin_value

    for index in range(coffin_index + 1, len(normalized_row)):
        if index > coffin_index:
            normalized_row[index] = ""

    return normalized_row


def shift_row_segment_right(row: list[str], start_index: int) -> list[str]:
    normalized_row = row[:]
    non_empty_indexes = [index for index, value in enumerate(normalized_row) if value.strip()]
    if not non_empty_indexes:
        return normalized_row

    last_non_empty_index = non_empty_indexes[-1]
    if last_non_empty_index >= len(normalized_row) - 1:
        return normalized_row

    for index in range(last_non_empty_index + 1, start_index, -1):
        normalized_row[index] = normalized_row[index - 1]
    normalized_row[start_index] = ""
    return normalized_row


def shift_misaligned_identity_block(headers: list[str], row: list[str]) -> list[str]:
    normalized_row = row[:]
    gender_index = find_header_index(headers, ("性別", "性"))
    age_index = find_header_index(headers, ("年齡", "年龄"))
    origin_index = find_header_index(headers, ("籍貫", "籍贯"))
    address_index = find_header_index(headers, ("住址",))
    illness_index = find_header_index(headers, ("病狀", "病状"))
    family_index = find_header_index(headers, ("認家", "认家"))
    death_index = find_header_index(headers, ("死亡日期",))

    required_indexes = (gender_index, age_index, origin_index, address_index, illness_index, family_index, death_index)
    if any(index is None for index in required_indexes):
        return normalized_row

    max_index = max(index for index in required_indexes if index is not None)
    if len(normalized_row) <= max_index:
        normalized_row.extend([""] * (max_index + 1 - len(normalized_row)))

    gender_value = normalized_row[gender_index].strip()
    age_value = normalized_row[age_index].strip()
    origin_value = normalized_row[origin_index].strip()
    address_value = normalized_row[address_index].strip()
    illness_value = normalized_row[illness_index].strip()
    family_value = normalized_row[family_index].strip()
    death_value = normalized_row[death_index].strip()

    if not is_age_marker_like(gender_value):
        return normalized_row
    if age_value and is_age_like(age_value):
        return normalized_row
    if address_value:
        return normalized_row
    if not family_value or not is_date_like(family_value):
        return normalized_row
    if death_value and is_date_like(death_value):
        return normalized_row

    return shift_row_segment_right(normalized_row, gender_index)


def shift_place_fields_right(headers: list[str], row: list[str]) -> list[str]:
    normalized_row = row[:]
    age_index = find_header_index(headers, ("年齡", "年龄"))
    origin_index = find_header_index(headers, ("籍貫", "籍贯"))
    address_index = find_header_index(headers, ("住址",))

    required_indexes = (age_index, origin_index, address_index)
    if any(index is None for index in required_indexes):
        return normalized_row

    max_index = max(index for index in required_indexes if index is not None)
    if len(normalized_row) <= max_index:
        normalized_row.extend([""] * (max_index + 1 - len(normalized_row)))

    age_value = normalized_row[age_index].strip()
    origin_value = normalized_row[origin_index].strip()
    address_value = normalized_row[address_index].strip()
    original_origin_value = origin_value

    age_and_origin = split_age_and_origin(age_value)
    if age_and_origin and not is_address_like(age_and_origin[1]) and not address_value:
        normalized_row[age_index] = age_and_origin[0]
        normalized_row[origin_index] = age_and_origin[1]
        if is_address_like(original_origin_value):
            normalized_row[address_index] = original_origin_value
        age_value = normalized_row[age_index].strip()
        origin_value = normalized_row[origin_index].strip()
        address_value = normalized_row[address_index].strip()

    if address_value:
        return normalized_row
    if not origin_value:
        return normalized_row
    if is_origin_like(origin_value):
        return normalized_row
    if not is_address_like(origin_value):
        return normalized_row
    if age_value and not is_age_like(age_value):
        return normalized_row

    moved_row = normalized_row[:]
    moved_row[address_index] = origin_value
    moved_row[origin_index] = ""
    return moved_row


def shift_age_origin_block_right(headers: list[str], row: list[str]) -> list[str]:
    normalized_row = row[:]
    gender_index = find_header_index(headers, ("性別", "性"))
    age_index = find_header_index(headers, ("年齡", "年龄"))
    origin_index = find_header_index(headers, ("籍貫", "籍贯"))

    required_indexes = (gender_index, age_index, origin_index)
    if any(index is None for index in required_indexes):
        return normalized_row

    max_index = max(index for index in required_indexes if index is not None)
    if len(normalized_row) <= max_index:
        normalized_row.extend([""] * (max_index + 1 - len(normalized_row)))

    gender_value = normalized_row[gender_index].strip()
    age_value = normalized_row[age_index].strip()
    origin_value = normalized_row[origin_index].strip()

    if not is_age_marker_like(gender_value):
        return normalized_row
    if not age_value or is_age_like(age_value):
        return normalized_row
    if origin_value:
        return normalized_row
    if not is_origin_like(age_value):
        return normalized_row

    normalized_row[gender_index] = ""
    normalized_row[age_index] = gender_value
    normalized_row[origin_index] = age_value
    return normalized_row


def shift_gender_age_only_block(headers: list[str], row: list[str]) -> list[str]:
    normalized_row = row[:]
    gender_index = find_header_index(headers, ("性別", "性"))
    age_index = find_header_index(headers, ("年齡", "年龄"))
    origin_index = find_header_index(headers, ("籍貫", "籍贯"))
    address_index = find_header_index(headers, ("住址",))

    required_indexes = (gender_index, age_index, origin_index, address_index)
    if any(index is None for index in required_indexes):
        return normalized_row

    max_index = max(index for index in required_indexes if index is not None)
    if len(normalized_row) <= max_index:
        normalized_row.extend([""] * (max_index + 1 - len(normalized_row)))

    gender_value = normalized_row[gender_index].strip()
    age_value = normalized_row[age_index].strip()
    origin_value = normalized_row[origin_index].strip()
    address_value = normalized_row[address_index].strip()

    if not is_age_marker_like(gender_value):
        return normalized_row
    if age_value or origin_value:
        return normalized_row
    if not address_value:
        return normalized_row

    normalized_row[gender_index] = ""
    normalized_row[age_index] = gender_value
    return normalized_row


def normalize_identity_columns(headers: list[str], row: list[str]) -> list[str]:
    normalized_row = row[:]
    name_index = find_header_index(headers, ("姓名", "名別", "名字"))
    gender_index = find_header_index(headers, ("性別", "性"))
    age_index = find_header_index(headers, ("年齡", "年龄"))
    origin_index = find_header_index(headers, ("籍貫", "籍贯"))

    if name_index is None or gender_index is None:
        return normalized_row

    max_index = max(index for index in (name_index, gender_index, age_index, origin_index) if index is not None)
    if max_index >= len(normalized_row):
        normalized_row.extend([""] * (max_index + 1 - len(normalized_row)))

    name_value = normalized_row[name_index].strip()
    gender_value = normalized_row[gender_index].strip()
    age_value = normalized_row[age_index].strip() if age_index is not None else ""

    has_gender_suffix = name_value.endswith(("男", "女")) and len(name_value) > 1

    if not gender_value and has_gender_suffix:
        normalized_row[name_index] = name_value[:-1].strip()
        normalized_row[gender_index] = name_value[-1]
        name_value = normalized_row[name_index]
        gender_value = normalized_row[gender_index]

    if age_index is not None:
        if is_age_fragment(gender_value) and age_value in {"歲", "岁"}:
            combined_age = f"{gender_value}{age_value}"
            if is_age_marker_like(combined_age):
                normalized_row[gender_index] = ""
                normalized_row[age_index] = combined_age
                gender_value = normalized_row[gender_index]
                age_value = normalized_row[age_index]

        split_gender_age = split_gender_and_age(gender_value)
        if split_gender_age and not age_value:
            normalized_row[gender_index], normalized_row[age_index] = split_gender_age
            gender_value = normalized_row[gender_index]
            age_value = normalized_row[age_index]

        if not age_value and is_age_marker_like(normalize_age_text(gender_value)):
            normalized_row[age_index] = normalize_age_text(gender_value)
            normalized_row[gender_index] = ""
            gender_value = normalized_row[gender_index]
            age_value = normalized_row[age_index]

        if has_gender_suffix and is_age_like(gender_value) and not age_value:
            normalized_row[name_index] = name_value[:-1].strip()
            normalized_row[gender_index] = name_value[-1]
            normalized_row[age_index] = gender_value
            name_value = normalized_row[name_index]
            gender_value = normalized_row[gender_index]
            age_value = normalized_row[age_index]

        if not age_value and is_age_like(gender_value):
            normalized_row[age_index] = gender_value
            normalized_row[gender_index] = ""
            age_value = normalized_row[age_index]
            gender_value = normalized_row[gender_index]

        if not gender_value and name_value.endswith(("男", "女")) and len(name_value) > 1:
            normalized_row[name_index] = name_value[:-1].strip()
            normalized_row[gender_index] = name_value[-1]

    if age_index is not None and origin_index is not None:
        age_value = normalized_row[age_index].strip()
        origin_value = normalized_row[origin_index].strip()
        age_and_origin = split_age_and_origin(age_value)
        if age_and_origin and not origin_value:
            normalized_row[age_index], normalized_row[origin_index] = age_and_origin

    return normalized_row


def shift_gender_age_origin_block_right(headers: list[str], row: list[str]) -> list[str]:
    normalized_row = row[:]
    gender_index = find_header_index(headers, ("性別", "性"))
    age_index = find_header_index(headers, ("年齡", "年龄"))
    origin_index = find_header_index(headers, ("籍貫", "籍贯"))
    address_index = find_header_index(headers, ("住址",))

    required_indexes = (gender_index, age_index, origin_index, address_index)
    if any(index is None for index in required_indexes):
        return normalized_row

    max_index = max(index for index in required_indexes if index is not None)
    if len(normalized_row) <= max_index:
        normalized_row.extend([""] * (max_index + 1 - len(normalized_row)))

    gender_value = normalized_row[gender_index].strip()
    age_value = normalized_row[age_index].strip()
    origin_value = normalized_row[origin_index].strip()
    address_value = normalized_row[address_index].strip()

    split_gender_age = split_gender_and_age(gender_value)
    age_marker_value = split_gender_age[1] if split_gender_age else gender_value

    if not is_age_marker_like(age_marker_value):
        return normalized_row
    if not age_value or not is_origin_like(age_value):
        return normalized_row
    if origin_value:
        return normalized_row
    if not address_value:
        return normalized_row

    shifted_row = shift_row_segment_right(normalized_row, age_index)
    if split_gender_age:
        shifted_row[gender_index], shifted_row[age_index] = split_gender_age
    else:
        shifted_row[gender_index] = ""
        shifted_row[age_index] = gender_value
    shifted_row[origin_index] = age_value
    return shifted_row


def shift_age_in_gender_to_origin(headers: list[str], row: list[str]) -> list[str]:
    normalized_row = row[:]
    gender_index = find_header_index(headers, ("性別", "性"))
    age_index = find_header_index(headers, ("年齡", "年龄"))
    origin_index = find_header_index(headers, ("籍貫", "籍贯"))
    address_index = find_header_index(headers, ("住址",))

    required_indexes = (gender_index, age_index, origin_index, address_index)
    if any(index is None for index in required_indexes):
        return normalized_row

    max_index = max(index for index in required_indexes if index is not None)
    if len(normalized_row) <= max_index:
        normalized_row.extend([""] * (max_index + 1 - len(normalized_row)))

    gender_value = normalized_row[gender_index].strip()
    age_value = normalized_row[age_index].strip()
    origin_value = normalized_row[origin_index].strip()
    address_value = normalized_row[address_index].strip()

    split_gender_age = split_gender_and_age(gender_value)
    age_marker_value = split_gender_age[1] if split_gender_age else normalize_age_text(gender_value)
    if not is_age_marker_like(age_marker_value):
        return normalized_row
    if not age_value:
        return normalized_row
    if is_age_like(age_value) or is_address_like(age_value):
        return normalized_row
    if origin_value:
        return normalized_row
    if not address_value:
        return normalized_row

    if split_gender_age:
        normalized_row[gender_index], normalized_row[age_index] = split_gender_age
    else:
        normalized_row[gender_index] = ""
        normalized_row[age_index] = age_marker_value
    normalized_row[origin_index] = age_value
    return normalized_row


def shift_front_identity_columns_right(headers: list[str], row: list[str]) -> list[str]:
    normalized_row = row[:]
    gender_index = find_header_index(headers, ("性別", "性"))
    age_index = find_header_index(headers, ("年齡", "年龄"))
    origin_index = find_header_index(headers, ("籍貫", "籍贯"))
    address_index = find_header_index(headers, ("住址",))
    illness_index = find_header_index(headers, ("病狀", "病状"))
    family_index = find_header_index(headers, ("認家", "认家"))
    death_index = find_header_index(headers, ("死亡日期",))

    required_indexes = (gender_index, age_index, origin_index, address_index, illness_index, family_index, death_index)
    if any(index is None for index in required_indexes):
        return normalized_row

    max_index = max(index for index in required_indexes if index is not None)
    if len(normalized_row) <= max_index:
        normalized_row.extend([""] * (max_index + 1 - len(normalized_row)))

    gender_value = normalized_row[gender_index].strip()
    age_value = normalized_row[age_index].strip()
    origin_value = normalized_row[origin_index].strip()
    address_value = normalized_row[address_index].strip()
    illness_value = normalized_row[illness_index].strip()
    family_value = normalized_row[family_index].strip()
    death_value = normalized_row[death_index].strip()

    age_marker_value = normalize_age_text(gender_value)
    if not is_age_marker_like(age_marker_value):
        return normalized_row
    if not age_value or is_age_like(age_value) or is_address_like(age_value):
        return normalized_row

    if origin_value and is_address_like(origin_value):
        normalized_row[gender_index] = ""
        normalized_row[age_index] = age_marker_value
        normalized_row[origin_index] = age_value
        normalized_row[address_index] = origin_value
        return normalized_row

    if (
        origin_value
        and address_value
        and not is_address_like(address_value)
        and not illness_value
        and not family_value
        and is_date_like(death_value)
    ):
        normalized_row[gender_index] = ""
        normalized_row[age_index] = age_marker_value
        normalized_row[origin_index] = age_value
        normalized_row[address_index] = origin_value
        normalized_row[family_index] = address_value
        return normalized_row

    return normalized_row


def shift_age_origin_address_triplet_right(headers: list[str], row: list[str]) -> list[str]:
    normalized_row = row[:]
    gender_index = find_header_index(headers, ("性別", "性"))
    age_index = find_header_index(headers, ("年齡", "年龄"))
    origin_index = find_header_index(headers, ("籍貫", "籍贯"))
    address_index = find_header_index(headers, ("住址",))
    illness_index = find_header_index(headers, ("病狀", "病状"))

    required_indexes = (gender_index, age_index, origin_index, address_index, illness_index)
    if any(index is None for index in required_indexes):
        return normalized_row

    max_index = max(index for index in required_indexes if index is not None)
    if len(normalized_row) <= max_index:
        normalized_row.extend([""] * (max_index + 1 - len(normalized_row)))

    gender_value = normalized_row[gender_index].strip()
    age_value = normalized_row[age_index].strip()
    origin_value = normalized_row[origin_index].strip()
    address_value = normalized_row[address_index].strip()
    illness_value = normalized_row[illness_index].strip()

    split_gender_age = split_gender_and_age(gender_value)
    age_marker_value = split_gender_age[1] if split_gender_age else normalize_age_text(gender_value)
    if not is_age_marker_like(age_marker_value):
        return normalized_row
    if not age_value or is_age_like(age_value) or is_address_like(age_value):
        return normalized_row
    if not origin_value or not is_address_like(origin_value):
        return normalized_row
    if address_value:
        return normalized_row
    if not illness_value:
        return normalized_row

    normalized_row[gender_index] = split_gender_age[0] if split_gender_age else ""
    normalized_row[age_index] = age_marker_value
    normalized_row[origin_index] = age_value
    normalized_row[address_index] = origin_value
    return normalized_row


def shift_split_gender_age_block_right(headers: list[str], row: list[str]) -> list[str]:
    normalized_row = row[:]
    gender_index = find_header_index(headers, ("性別", "性"))
    age_index = find_header_index(headers, ("年齡", "年龄"))
    origin_index = find_header_index(headers, ("籍貫", "籍贯"))
    address_index = find_header_index(headers, ("住址",))
    illness_index = find_header_index(headers, ("病狀", "病状"))
    family_index = find_header_index(headers, ("認家", "认家"))
    death_index = find_header_index(headers, ("死亡日期",))
    grave_index = find_header_index(headers, ("墓地號數", "墓地号数"))
    coffin_index = find_header_index(headers, ("棺木", "類別", "类别"))

    required_indexes = (gender_index, age_index, origin_index, address_index, illness_index, family_index, death_index, grave_index, coffin_index)
    if any(index is None for index in required_indexes):
        return normalized_row

    max_index = max(index for index in required_indexes if index is not None)
    if len(normalized_row) <= max_index:
        normalized_row.extend([""] * (max_index + 1 - len(normalized_row)))

    gender_value = normalized_row[gender_index].strip()
    age_value = normalized_row[age_index].strip()
    origin_value = normalized_row[origin_index].strip()
    address_value = normalized_row[address_index].strip()
    illness_value = normalized_row[illness_index].strip()
    family_value = normalized_row[family_index].strip()
    death_value = normalized_row[death_index].strip()
    grave_value = normalized_row[grave_index].strip()

    split_gender_age = split_gender_and_age(gender_value)
    if not split_gender_age:
        return normalized_row
    if not is_origin_like(age_value):
        return normalized_row
    if not origin_value or not is_address_like(origin_value):
        return normalized_row
    if address_value:
        return normalized_row
    if not illness_value:
        return normalized_row
    if not is_date_like(family_value):
        return normalized_row
    if not death_value or not grave_value:
        return normalized_row

    shifted_row = shift_row_segment_right(normalized_row, age_index)
    shifted_row[gender_index], shifted_row[age_index] = split_gender_age
    return shifted_row


def shift_full_identity_chain_right(headers: list[str], row: list[str]) -> list[str]:
    normalized_row = row[:]
    gender_index = find_header_index(headers, ("性別", "性"))
    age_index = find_header_index(headers, ("年齡", "年龄"))
    origin_index = find_header_index(headers, ("籍貫", "籍贯"))
    address_index = find_header_index(headers, ("住址",))
    illness_index = find_header_index(headers, ("病狀", "病状"))
    family_index = find_header_index(headers, ("認家", "认家"))
    death_index = find_header_index(headers, ("死亡日期",))
    grave_index = find_header_index(headers, ("墓地號數", "墓地号数"))
    coffin_index = find_header_index(headers, ("棺木", "類別", "类别"))

    required_indexes = (
        gender_index,
        age_index,
        origin_index,
        address_index,
        illness_index,
        family_index,
        death_index,
        grave_index,
        coffin_index,
    )
    if any(index is None for index in required_indexes):
        return normalized_row

    max_index = max(index for index in required_indexes if index is not None)
    if len(normalized_row) <= max_index:
        normalized_row.extend([""] * (max_index + 1 - len(normalized_row)))

    gender_value = normalized_row[gender_index].strip()
    age_value = normalized_row[age_index].strip()
    origin_value = normalized_row[origin_index].strip()
    address_value = normalized_row[address_index].strip()
    illness_value = normalized_row[illness_index].strip()
    family_value = normalized_row[family_index].strip()
    death_value = normalized_row[death_index].strip()
    grave_value = normalized_row[grave_index].strip()

    age_marker_value = normalize_age_text(gender_value)
    if not is_age_marker_like(age_marker_value):
        return normalized_row
    if not age_value or is_age_like(age_value) or is_address_like(age_value):
        return normalized_row
    if not origin_value or not is_address_like(origin_value):
        return normalized_row
    if address_value:
        return normalized_row
    if not illness_value:
        return normalized_row
    if not family_value or not is_date_like(family_value):
        return normalized_row
    if not death_value or not grave_value:
        return normalized_row

    shifted_row = shift_row_segment_right(normalized_row, gender_index)
    shifted_row[gender_index] = ""
    shifted_row[age_index] = age_marker_value
    return shifted_row


def shift_tail_fields_right(headers: list[str], row: list[str]) -> list[str]:
    normalized_row = row[:]
    illness_index = find_header_index(headers, ("病狀", "病状"))
    family_index = find_header_index(headers, ("認家", "认家"))
    death_index = find_header_index(headers, ("死亡日期",))
    grave_index = find_header_index(headers, ("墓地號數", "墓地号数"))
    coffin_index = find_header_index(headers, ("棺木", "類別", "类别"))

    required_indexes = (illness_index, family_index, death_index, grave_index, coffin_index)
    if any(index is None for index in required_indexes):
        return normalized_row

    max_index = max(index for index in required_indexes if index is not None)
    if len(normalized_row) <= max_index:
        normalized_row.extend([""] * (max_index + 1 - len(normalized_row)))

    illness_value = normalized_row[illness_index].strip()
    family_value = normalized_row[family_index].strip()
    death_value = normalized_row[death_index].strip()
    grave_value = normalized_row[grave_index].strip()
    coffin_value = normalized_row[coffin_index].strip()

    if not illness_value:
        return normalized_row
    if not family_value or not is_date_like(family_value):
        return normalized_row
    if not death_value:
        return normalized_row
    if not grave_value:
        return normalized_row
    if coffin_value:
        return normalized_row

    normalized_row[coffin_index] = grave_value
    normalized_row[grave_index] = death_value
    normalized_row[death_index] = family_value
    normalized_row[family_index] = illness_value
    normalized_row[illness_index] = ""
    return normalized_row


def normalize_row_break_separators(headers: list[str], rows: list[list[str]]) -> list[list[str]]:
    expected_width = len(headers)
    normalized_rows: list[list[str]] = []
    for row in rows:
        normalized_row = split_br_delimited_cells(row, expected_width)
        normalized_row = normalize_identity_columns(headers, normalized_row)
        normalized_row = shift_age_in_gender_to_origin(headers, normalized_row)
        normalized_row = shift_age_origin_address_triplet_right(headers, normalized_row)
        normalized_row = shift_front_identity_columns_right(headers, normalized_row)
        normalized_row = shift_full_identity_chain_right(headers, normalized_row)
        normalized_row = shift_tail_fields_right(headers, normalized_row)
        normalized_row = shift_split_gender_age_block_right(headers, normalized_row)
        normalized_row = shift_gender_age_origin_block_right(headers, normalized_row)
        normalized_row = shift_age_origin_block_right(headers, normalized_row)
        normalized_row = shift_gender_age_only_block(headers, normalized_row)
        normalized_row = shift_misaligned_identity_block(headers, normalized_row)
        normalized_row = shift_place_fields_right(headers, normalized_row)
        normalized_row = normalize_compact_obituary_row(headers, normalized_row)
        normalized_row = normalize_identity_columns(headers, normalized_row)
        normalized_rows.append(normalized_row)
    return normalized_rows


def sanitize_sheet_name(name: str, used_names: set[str]) -> str:
    cleaned = re.sub(r'[\\/*?:\[\]]', "_", name).strip() or "Sheet"
    cleaned = cleaned[:31]
    candidate = cleaned
    counter = 1

    while candidate in used_names:
        suffix = f"_{counter}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        counter += 1

    used_names.add(candidate)
    return candidate


def build_output_path(base_dir: Path, preferred_name: str = "markdown_tables.xlsx") -> Path:
    output_path = base_dir / preferred_name
    if not output_path.exists():
        return output_path

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return base_dir / f"{output_path.stem}_{timestamp}{output_path.suffix}"


def write_table_to_sheet(worksheet, start_row: int, table: TableBlock) -> int:
    current_row = start_row

    worksheet.cell(row=current_row, column=1, value=table.title)
    worksheet.cell(row=current_row, column=1).font = Font(bold=True, size=12)
    current_row += 1

    for column_index, header in enumerate(table.headers, start=1):
        cell = worksheet.cell(row=current_row, column=column_index, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=False)
    current_row += 1

    for row in table.rows:
        padded_row = row + [""] * (len(table.headers) - len(row))
        for column_index, value in enumerate(padded_row, start=1):
            cell = worksheet.cell(row=current_row, column=column_index, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=False)
        current_row += 1

    worksheet.cell(row=current_row, column=1, value=f"表格编号：{table.table_number}")
    worksheet.cell(row=current_row, column=1).font = Font(italic=True)
    current_row += 1

    return current_row + 4


def auto_adjust_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 40)


def export_tables_to_workbook(parsed_files: list[tuple[Path, list[TableBlock]]], output_path: str | Path) -> str:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    used_names: set[str] = set()

    for file_path, tables in parsed_files:
        sheet_name = sanitize_sheet_name(file_path.stem, used_names)
        worksheet = workbook.create_sheet(title=sheet_name)

        if not tables:
            worksheet.cell(row=1, column=1, value=f"{file_path.name} 中未找到 Markdown 表格。")
            worksheet.cell(row=1, column=1).font = Font(bold=True)
            auto_adjust_columns(worksheet)
            continue

        current_row = 1
        for table in tables:
            table.rows = normalize_row_break_separators(table.headers, table.rows)
            current_row = write_table_to_sheet(worksheet, current_row, table)

        auto_adjust_columns(worksheet)
    workbook.save(output_path)
    return str(output_path)


def select_markdown_files() -> list[str]:
    import tkinter as tk
    from tkinter import filedialog

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    file_paths = filedialog.askopenfilenames(
        title="选择要导出的 Markdown 文件",
        filetypes=[("Markdown 文件", "*.md"), ("所有文件", "*.*")],
    )
    root.destroy()
    return list(file_paths)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="将 Markdown 表格批量导出到 Excel。")
    parser.add_argument("files", nargs="*", help="可选：直接传入一个或多个 Markdown 文件路径。")
    parser.add_argument(
        "--output",
        help="可选：指定输出 Excel 文件路径；不传时默认输出到项目目录。",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    selected_files = [str(Path(file).resolve()) for file in args.files] if args.files else select_markdown_files()

    if not selected_files:
        print("未选择任何 Markdown 文件，程序已退出。")
        return 0

    parsed_files: list[tuple[Path, list[TableBlock]]] = []
    for file_name in selected_files:
        file_path = Path(file_name)
        parsed_files.append((file_path, parse_markdown_tables(file_path)))

    project_dir = Path(__file__).resolve().parent
    output_path = Path(args.output).resolve() if args.output else build_output_path(project_dir)
    result = export_tables_to_workbook(parsed_files, output_path)
    print(f"导出完成：{result}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
